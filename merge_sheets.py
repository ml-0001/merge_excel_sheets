# coding=UTF-8
from __future__ import unicode_literals
import sys,os
reload(sys)
sys.setdefaultencoding('gb2312')
import logging
import openpyxl


def usage():
    print("Usage:%s -d <merged excel file name> <to be merged excel file list>")
    print("Example:%s test_merged.xlsx test01.xlsx test02.xlsx")
    print("Example:%s test_merged.xlsx -dir <to be merged excel directory name>")


def merge_excel_sheets(merged_file, tobe_merged):
    #using 1st excel format as standard
    std_fmt_file =  tobe_merged[0]
    std_fmt_data = openpyxl.load_workbook(std_fmt_file)
    sheet_names =  std_fmt_data.sheetnames
    #
    logging.debug(u'merge file list:"%s" to a single excel:"%s"', u','.join(tobe_merged), merged_file)
    
    #initialize
    wb_merged_dict = dict()
    for name in sheet_names:
        wb_merged_dict[name] = [[]]
        std_sheet = std_fmt_data[name]
        #first line (row) is table head
        for head_col_idx in range(std_sheet.max_column):
            wb_merged_dict[name][0].append(std_sheet.cell(row=1,column=head_col_idx+1).value)

    #merge all excel by sheet 
    for excel in tobe_merged:
        logging.debug('merging file:"%s" ...', excel)
        ws_data=openpyxl.load_workbook(excel)
        for sheet_name in sheet_names:
            if sheet_name not in ws_data:
                logging.warn('WARNING: excel file: "%s" has no sheet named: "%s" will ignore it .', excel, sheet_name)
                continue
            sheet = ws_data[sheet_name]
            sheet_row = 0
            for row_idx in range(sheet.max_row):
                if row_idx == 0:
                    continue
                sheet_row = sheet_row + 1
                row_data_list = [] 
                for col_idx in range(sheet.max_column):
                    row_data_list.append(sheet.cell(row=row_idx+1, column=col_idx+1).value)
                wb_merged_dict[sheet_name].append(row_data_list) 
            ########################################################################
            logging.debug('merged excel:"%s" sheet:"%s" row num:%d/%d', excel, sheet_name, sheet_row, sheet.max_row)

    #save data without format  (style)
    wb_merged = openpyxl.Workbook() #xlwt.Workbook(encoding = 'UTF-8')
    def_sheet = wb_merged.active
    wb_merged.remove(def_sheet)
    logging.debug("merged all file success , saving file:%s ...", merged_file)
    for sheet_name,sheet_data in wb_merged_dict.items():
        sheet = wb_merged.create_sheet(title=sheet_name) # wb_merged.add_sheet(sheet_name)
        for rowx in range(len(sheet_data)):
            for colx in range(len(sheet_data[rowx])):
                sheet.cell(row=rowx+1, column=colx+1, value=sheet_data[rowx][colx])
        logging.info('merged sheet:"%s" all files with data total row:%d .', sheet_name, len(sheet_data))
    wb_merged.save(merged_file)

def main():
    if len(sys.argv) < 3:
        usage()
        os.exit(2)
    merged = sys.argv[1]
    logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s', level=logging.DEBUG)
    if sys.argv[2] == '-dir':
        if len(sys.argv) < 4:
            print("error command format , expect directory name after -dir")
            usage()
            os.exit(3)
        dirname = sys.argv[3]
        tobe_merged = []
        for dn in os.listdir(dirname):
            if dn.endswith('.xlsx'):
                tobe_merged.append(os.path.join(dirname, dn))
        
        if len(tobe_merged) == 0:
            logging.error("not found xlsx file in dir:%s", dirname)
    else:
        tobe_merged = sys.argv[2:]
    ########################## 
    merge_excel_sheets(merged, tobe_merged)


####################################################
main()
